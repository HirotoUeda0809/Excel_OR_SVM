import cvxopt as cvx
import numpy as np
import openpyxl as px
import os

path = os.path.expanduser("~")

Data5 = px.load_workbook("{:}\\Desktop\\seminar\\Excel_OR\\pycode\\Chap5.xlsx".format(path))
data_fuji = Data5["例題5.3.2"]

DataNum = 10
ColumnNum = 30

def y_vector(data_num):
    y = [] #B5セルから読み込み
    for i in range(0, data_num):
        if (data_fuji.cell(row = i + 5, column = 2).value == "M"):
            y += [1]
        else:
            y += [-1]
    return y

y = y_vector(DataNum)

#目的関数部分の行列生成
#1次式部分の係数読み込み(SVMは2次形式のものしかないのでここは関係ない)
c_degree1 = np.array([0 for i in range(0, ColumnNum + 1)]).astype(np.float)
#2次式部分の係数読み込み(k = katamuki, s= seppen)
c_d2_k = np.diag(np.array([1 for i in range(1, ColumnNum + 1 + 1)]))
c_d2_s = np.array([0 for i in range(1, ColumnNum + 1 + 1)])
c_d2_k[:, ColumnNum] = c_d2_s
c_degree2 = c_d2_k.astype(np.float)
#cvxoptの型に変換
q = cvx.matrix(c_degree1)
P = cvx.matrix(c_degree2)

#制約条件部の行列生成(-1のスカラー倍は制約を<=から>=にするためで，SVMは>=であるが，cvxoptは<=しか許さない)
#左辺Gx
#本制約部(C4セルから読み込み)
c_constraint_l_k = np.array([[y[i] * data_fuji.cell(row = i + 5, column = j + 3).value for j in range(0, ColumnNum)] for i in range(0, DataNum)])
c_constraint_l_s = np.array([[y[i]] for i in range(0, DataNum)])
c_constraint_l = -1 * np.hstack((c_constraint_l_k, c_constraint_l_s)).astype(np.float)
#統合して1つの行列を生成し，cvxoptの型に変換
G = cvx.matrix(c_constraint_l)
print(G)

#右辺h
c_constraint_r = -1 * np.array([1 for i in range(1, DataNum + 1)]).astype(np.float)
h = cvx.matrix(c_constraint_r)

#ソルバー起動と解表示
#-Gx <= -h <=> Gx >= h
cvx.solvers.options["feastol"] = 1e-12
solve = cvx.solvers.qp(P = P, q = q, G = G, h = h)
print(solve)
print(solve["x"])
#print(solve["x"][0])
print(solve["primal objective"])
