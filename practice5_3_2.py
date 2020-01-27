#################################################################################
# 作成日: 2020/01/26 
# Support Vector Machine(凸2次計画問題)をCVXOPTというPythonのパッケージを用いて解く 
#################################################################################

#ライブラリ読み込み
import cvxopt as cvx
import numpy as np
import openpyxl as px
import os
import mosek

#ExcelファイルまでのPATH
path = os.path.expanduser("~")
Data5 = px.load_workbook("{:}\\Desktop\\seminar\\Excel_OR\\pycode\\Chap5.xlsx".format(path))
data_name = Data5["演習5.3.1"]
#data_name = Data5["例題5.3.2"]

#目的関数のうち2次形式部分の係数行列を生成し，cvx.matrix形式で返す．
def P_matrix(data_num, column_num):
    vk = 2 * column_num
    v = 2 + vk
    P_matrix = []
    for i in range(0, data_num + v):
        P_matrix_row = []
        for j in range(0, data_num + v):
            if (j in range(0, vk) and j == i):
                P_matrix_row += [1]
            elif (j in range(0, vk) and i % 2 == 0 and j == i + 1):
                P_matrix_row += [-1]
            elif (j in range(0, vk) and i % 2 != 0 and j == i - 1):
                P_matrix_row += [-1]
            else:
                P_matrix_row += [0]
        P_matrix += [P_matrix_row]
    P_matrix = cvx.matrix(np.array(P_matrix).astype(np.float))
    return P_matrix
#目的関数のうち1次式部分の係数ベクトルを生成し，cvx.matrix形式で返す．
def q_vector(data_num, column_num, param):
    v = 2 * column_num + 2
    c_degree1_soft = param * np.array([0 if i in range(0, v) else 1 for i in range(0, data_num + v)])
    q_vector = cvx.matrix(c_degree1_soft.astype(np.float))
    return q_vector
#制約条件(左辺)の係数行列(G)を生成し，cvx.matrix形式で返す (ただしGx >= h)．
def G_matrix(data_num, column_num, start_row, start_col, label):
    #ラベルの値によって+1か-1を割り当てる
    def y_vector(data_num, label):
        y = [] #B4セルから読み込み
        for i in range(0, data_num):
            if (data_name.cell(row = i + start_row, column = start_col - 1).value == label):
                y += [1]
            else:
                y += [-1]
        return y

    #SVMのうち主たる制約の係数行列を生成する関数
    def Primal_Const(data_num, column_num, start_row, start_col, label):
        def G_k(data_num, column_num, start_row, start_col, label):
            y = y_vector(data_num, label)
            sr = start_row; sc = start_col
            c_constraint_k = []
            for i in range(0, data_num):
                c_constraint_k_row = []
                for j in range(0, column_num):
                    c_constraint_k_row += [y[i] * data_name.cell(row = i + sr, column = j + sc).value, -y[i] * data_name.cell(row = i + sr, column = j + sc).value]
                c_constraint_k += [c_constraint_k_row]
            c_constraint_k = np.array(c_constraint_k)
            return c_constraint_k
        def G_s(data_num, label):
            y = y_vector(data_num, label)
            c_constraint_s = []
            for i in range(0, data_num):
                c_constraint_s_row = []
                c_constraint_s_row += [y[i], -y[i]]
                c_constraint_s += [c_constraint_s_row]
            c_constraint_s = np.array(c_constraint_s)
            return c_constraint_s
        def G_xi(data_num):
            G_xi = np.diag(np.array([1 for i in range(0, data_num)]))
            return G_xi
        G_k = G_k(data_num, column_num, start_row, start_col, label)
        G_s = G_s(data_num, label)
        G_xi = G_xi(data_num)
        Primal_Const = np.hstack((G_k, G_s, G_xi))
        return Primal_Const
    #2つの変数に分割したのでそれぞれに非負制約をかける
    def Non_Negative_Const(data_num, column_num):
        v = 2 * column_num + 2
        non_nega = np.diag(np.array([1 for i in range(0, data_num + v)]))
        return non_nega
    PC = Primal_Const(data_num, column_num, start_row, start_col, label)
    NNC = Non_Negative_Const(data_num, column_num)
    G_matrix_soft = cvx.matrix(-1 * np.vstack((PC, NNC)).astype(np.float))
    return G_matrix_soft
#制約条件(右辺)の係数ベクトルを生成し，cvx.matrix形式で返す (ただしGx >= h)．
def h_vector(data_num, column_num):
    v = 2 * column_num + 2
    c_constraint_r =  np.array([1 for i in range(0, data_num)])
    non_negative_r = np.array([0 for i in range(0, data_num + v)])
    h = np.hstack((c_constraint_r, non_negative_r))
    h_vector = cvx.matrix(-1 * h.astype(np.float))
    return h_vector

#定式化
#データ数と特徴ベクトルの個数
DataNum = 50
ColumnNum = 2 #特徴の数
ExcelStartRow = 4 #例題5.3.2 => 5
ExcelStartCol = 3
Label = "継続" #例題5.3.2 => "M"
param = 10 #トレードオフパラメータ

#ベクトルと行列の生成
P = P_matrix(DataNum, ColumnNum)
q = q_vector(DataNum, ColumnNum, param)
G = G_matrix(DataNum, ColumnNum, ExcelStartRow, ExcelStartCol, Label)
h = h_vector(DataNum, ColumnNum)

#ソルバー起動と解表示
#cvx.solvers.options["maxiters"]
cvx.solvers.options["feastol"] = 1e-12 #他に ["abstol"], ["reltol"]がある 
cvx.solvers.options["mosek"] = {mosek.iparam.log: 0} #solver MOSEKで解く場合

solve = cvx.solvers.qp(P = P, q = q, G = G, h = h)

print(solve["status"])
print(solve["x"])
#print("{:.10f}".format(solve["x"][0] - solve["x"][1]))
print(solve["primal objective"])